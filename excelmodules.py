import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Side
import win32com.client
import pandas as pd
from openpyxl.utils import column_index_from_string
import os   # pdf printing....


def create_df_for_testing(drug_count):
    date_col = ['PATDOB']
    df = pd.read_csv('database_for_testing.csv', dtype={'PATZIP': str,
                                                        'DOCZIP': str}, parse_dates=date_col)
    return df.head(drug_count)


def convert_to_pdf(excelFilePath,pdfOutFilePath):
    try:
        xlApp = win32com.client.Dispatch("Excel.Application")
        xlApp.Visible = False

        wb = xlApp.Workbooks.Open(excelFilePath)

        ws = wb.Worksheets[0]
        ws.PageSetup.Zoom = False

        ws.PageSetup.FitToPagesTall = 1

        ws.PageSetup.FitToPagesWide = 1

        xlTypePDF = 0
        xlQualityStandard = 0

        ws.ExportAsFixedFormat(xlTypePDF,
                               pdfOutFilePath,
                               xlQualityStandard, True, True)
    except Exception as e:
        print(e)

    finally:
        wb.Close(False)
        xlApp.Quit
        wb = None
        xlApp = None


def add_logo(ws):
    # A wrapper over PIL.Image, used to provide image
    # inclusion properties to openpyxl library
    img = openpyxl.drawing.image.Image('HelloRx Final.png')
    img.width = 220
    img.height = 90

    # The Coordinates where the image would be pasted
    # (an image could span several rows and columns
    # depending on it's size)
    img.anchor = 'B1'

    # Adding the image to the worksheet
    # (with attributes like position)
    ws.add_image(img)


def find_last_row_with_text(sheet, column_to_find_last_row):
    for row in range(sheet.max_row, 0, -1):
        if sheet[column_to_find_last_row + str(row)].value is not None:
            return row
    return None


def get_cell_key_coord(ws, key_value):
    # Iterate over all cells in the sheet
    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell value matches the search value
            if key_value in str(cell.value):
                # If a match is found, print the cell address
                return cell.coordinate


def set_form_title(ws, form_title_address, form_title):
    # SET TITLE
    ws[form_title_address].value = form_title
    ws[form_title_address].font = openpyxl.styles.Font(size=16, bold=True)
    ws.row_dimensions[ws[form_title_address].row].height = 30


def set_form_note(ws, form_note_address, form_note):
    ws[form_note_address].value = form_note


def set_confidential_note(ws, confidential_notice_address, last_column):
    confidentiality_message = 'CONFIDENTIALITY NOTICE: The documents accompanying this facsimile transmittal are intended only ' \
                              'for the use of the individual or entity to which it is addressed. It may contain information ' \
                              'that is privileged, confidential and exempt from disclosure under law. If the reader of ' \
                              'this message is not the intended recipient, you are notified that any dissemination, ' \
                              'distribution or copying of this communication is strictly prohibited. If you are not the ' \
                              'intended recipient, you are hereby notified that law strictly prohibits any disclosure, ' \
                              'copying, distribution or action taken in reliance on the contents of these documents. ' \
                              'If you have received this fax in error, please notify the sender immediately to arrange for' \
                              ' return of these documents.'

    # SET CONFIDENTIALITY SECTION
    # Define the range to be merged
    confidentiality_row = ws[confidential_notice_address].row
    confidentiality_message_cell_start = 'A' + str(confidentiality_row)
    confidentiality_message_cell_end = ':' + last_column + str(confidentiality_row)
    confidentiality_message_cell_range = confidentiality_message_cell_start + confidentiality_message_cell_end

    # Merge the cells
    ws.merge_cells(confidentiality_message_cell_range)

    # Write confidentiality_message to the merged cell
    ws[confidentiality_message_cell_start] = confidentiality_message

    # Set word wrap for confidentiality_message_cell
    ws[confidentiality_message_cell_start].alignment = openpyxl.styles.Alignment(wrap_text=True,
                                                                                 horizontal='center',
                                                                                 vertical='distributed',
                                                                                 shrinkToFit=True)

    ws[confidentiality_message_cell_start].font = openpyxl.styles.Font(size=9, bold=True)

    # set height of confidentiality_message row
    ws.row_dimensions[confidentiality_row].height = 65


def draw_each_drug_border(ws,cell_row, cell_col,last_column):
    # Options = medium ; thin
    medium_line = Side(border_style='thin')
    thin_line = Side(border_style='thin')

    index_col = cell_col - 1
    label_col = cell_col
    drug_info_col = cell_col + 1

    top_row = cell_row
    bottom_row = cell_row + 1

    indexBorder = Border(top=medium_line,bottom=medium_line,left=medium_line,right=thin_line)
    topLabelBorder = Border(top=medium_line,bottom=thin_line,left=medium_line,right=thin_line)
    botLabelBorder = Border(top=thin_line, bottom=medium_line, left=medium_line, right=thin_line)
    topDrugInfoBorder = Border(top=medium_line,bottom=thin_line,left=medium_line,right=medium_line)
    botDrugInfoBorder = Border(top=thin_line, bottom=medium_line, left=medium_line, right=medium_line)

    # SET INDEX BORDER
    ws.cell(cell_row, index_col).border = indexBorder
    ws.cell(cell_row+1, index_col).border = indexBorder

    # SET LABEL BORDER
    ws.cell(top_row, label_col).border = topLabelBorder
    ws.cell(bottom_row, label_col).border = botLabelBorder

    # SET DRUG BORDER
    start_col_idx = drug_info_col
    end_col_idx = column_index_from_string(last_column)

    # Loop through the columns in the merged cell (Drug Info) and add the border
    for col_idx in range(start_col_idx, end_col_idx + 1):
        ws.cell(top_row, col_idx).border = topDrugInfoBorder
        ws.cell(bottom_row, col_idx).border = botDrugInfoBorder


def adjust_each_drug_cell_height(ws,cell_row, cell_col, max_char_one_line,height_increase_increment):
    drug_info_col = cell_col + 1
    bottom_row = cell_row + 1

    # Get the cell object
    sig_cell = ws.cell(row=bottom_row, column=drug_info_col)

    # Get the contents of the cell
    contents = sig_cell.value

    # Calculate the number of characters in the cell contents
    char_count = len(contents) if contents is not None else 0

    row_height = (char_count // max_char_one_line + 1) * height_increase_increment        # // = whole number

    ws.row_dimensions[bottom_row].height = row_height

def style_each_drug_info(ws, cell_row, cell_col, drug_info_top, drug_info_bottom, count):
    index_col = cell_col - 1
    label_col = cell_col
    drug_info_col = cell_col + 1

    top_row = cell_row
    bottom_row = cell_row + 1

    # SET DRUG Format
    ws.cell(bottom_row, drug_info_col).font = openpyxl.styles.Font(size=11, bold=False)


def set_each_drug_info(ws, cell_row, cell_col, drug_info_top, drug_info_bottom, count):
    index_col = cell_col - 1
    label_col = cell_col
    drug_info_col = cell_col + 1

    top_row = cell_row
    bottom_row = cell_row + 1

    # SET INDEX
    ws.cell(cell_row, index_col).value = count

    # SET LABEL
    ws.cell(top_row, label_col).value = 'Drug'
    # ws.cell(bottom_row, label_col).value = 'Sig'

    # SET DRUG INFO
    ws.cell(top_row, drug_info_col).value = drug_info_top
    ws.cell(bottom_row, drug_info_col).value = drug_info_bottom


def to_do_when_df_empty(ws,first_drug_cell_coord,last_column):
    drug_cell_row = ws[first_drug_cell_coord].row
    drug_cell_col = ws[first_drug_cell_coord].column

    # SET LABEL
    ws.cell(drug_cell_row, drug_cell_col).value = ''
    ws[last_column + str(ws[first_drug_cell_coord].row - 1)].value = ''

    # Draw Borders around Note Section
    medium_line = Side(border_style='medium')
    borderLine = Border(top=medium_line, bottom=medium_line, left=medium_line, right=medium_line)

    # Loop through the columns in the merged cell (Drug Info) and add the border
    start_col_idx = drug_cell_col
    end_col_idx = column_index_from_string(last_column)
    for col_idx in range(start_col_idx, end_col_idx + 1):
        ws.cell(drug_cell_row-2, col_idx).border = borderLine

    # Format Note Section -------------------------

def set_signature_date_section(ws, df,signature_cell_address):
    signature_col = ws[signature_cell_address].column
    signature_row = ws[signature_cell_address].row

    if df.empty:
        ws.cell(signature_row, signature_col).value = ''
    else:
        ws.cell(signature_row,signature_col ).value = 'Substitution Allowed: Yes____ No____     Signature: _____________        Date: _____________ '

        # Style
        ws.row_dimensions[signature_row].height = 20
        ws.cell(signature_row,signature_col).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='bottom', shrinkToFit=True, horizontal='center')
        ws.cell(signature_row,signature_col).font = openpyxl.styles.Font(size=14, bold=True)

def set_drugs(df, ws, first_drug_cell_coord, last_column, max_char_one_line):
    drug_cell_row = ws[first_drug_cell_coord].row
    drug_cell_col = ws[first_drug_cell_coord].column

    # iterate over rows
    index = 1
    for value in df[['drug_top_line_info', 'Direction']].values:
        drug_top_line_info = value[0]
        drug_bottom_line_info = value[1]

        set_each_drug_info(ws=ws, cell_row=drug_cell_row, cell_col=drug_cell_col,
                           drug_info_top=drug_top_line_info, drug_info_bottom=drug_bottom_line_info, count=index)

        style_each_drug_info(ws=ws, cell_row=drug_cell_row, cell_col=drug_cell_col,
                           drug_info_top=drug_top_line_info, drug_info_bottom=drug_bottom_line_info, count=index)

        draw_each_drug_border(ws=ws, cell_row=drug_cell_row, cell_col=drug_cell_col, last_column=last_column)

        adjust_each_drug_cell_height(ws=ws, cell_row=drug_cell_row, cell_col=drug_cell_col,
                                     max_char_one_line=max_char_one_line, height_increase_increment=14)

        index = index + 1
        drug_cell_row = drug_cell_row + 2


def hide_extra_row(ws,last_row,confidential_row):
    for i in range (last_row+2, confidential_row-2):
        ws.row_dimensions[i].hidden = True

def adjust_padding_row_height(ws, drug_count,confidential_coord):
    # set height of confidentiality_message row
    padding_row = ws[confidential_coord].row - 1
    drug_count = 1 if drug_count < 1 else drug_count
    ws.row_dimensions[padding_row].height = 2 * pow((15-drug_count),2)       # square power = more space when less drug_count

def set_approve_title(ws,drug_begin_address,last_column):
    ws[last_column + str(ws[drug_begin_address].row - 1)].value = 'Approve'

def open_and_get_worksheet_key(excelTemplateFilePath):
    wb = openpyxl.load_workbook(filename=excelTemplateFilePath)
    ws = wb.active

    form_title_key = 'form_title'
    time_stamp_key = 'request_time_stamp'
    patient_info_key = 'patlname'
    provider_info_key = 'providerlname'
    form_note_key = 'note_message'
    drug_begin_key = 'drug_key'
    confidential_notice_key = 'confidential_notice'
    signature_key = 'signature_key'

    form_title_address = get_cell_key_coord(ws, key_value=form_title_key)
    time_stamp_address = get_cell_key_coord(ws, key_value=time_stamp_key)
    patient_info_address = get_cell_key_coord(ws, key_value=patient_info_key)
    provider_info_address = get_cell_key_coord(ws, key_value=provider_info_key)
    form_note_address = get_cell_key_coord(ws, key_value=form_note_key)
    drug_begin_address = get_cell_key_coord(ws, key_value=drug_begin_key)
    confidential_notice_address = get_cell_key_coord(ws, key_value=confidential_notice_key)
    signature_key_address = get_cell_key_coord(ws, key_value=signature_key)

    key_coord_dict = {'form_title_address': form_title_address,
                'time_stamp_address': time_stamp_address,
                'patient_info_address': patient_info_address,
                'provider_info_address': provider_info_address,
                'form_note_address': form_note_address,
                'drug_begin_address': drug_begin_address,
                'confidential_notice_address': confidential_notice_address,
                'signature_key_address': signature_key_address}

    return wb, ws, key_coord_dict

def find_last_row(ws,col_to_find_last_row):
    last_row_1 = find_last_row_with_text(sheet=ws,
                                      column_to_find_last_row=col_to_find_last_row[0])

    last_row_2 = find_last_row_with_text(sheet=ws,
                                      column_to_find_last_row=col_to_find_last_row[1])

    return max(last_row_1,last_row_2)

def transfer_df_to_drug_table(ws,df,key_coord_dict,last_column,max_char_one_line):
    # To do if df_drug is empty
    if df.empty:
        to_do_when_df_empty(ws, first_drug_cell_coord=key_coord_dict['drug_begin_address'],
                            last_column=last_column)
    else:
        set_drugs(df=df, ws=ws,
                  first_drug_cell_coord=key_coord_dict['drug_begin_address'],
                  last_column=last_column,
                  max_char_one_line=max_char_one_line)

def set_patient_info(ws, key_coord_dict,window):
    pass


def set_recipient_info(ws, key_coord_dict,window):
    pass

def set_request_date_time(ws, date_time_cell_address):
    ws[date_time_cell_address].value = datetime.datetime.now().strftime("Request Time: %b %d, %Y at %H:%M %p")

def thank_you_note_checker(message):
    if not "thank" in message.lower():
        return message + " Thank you!"
    else:
        return message

def processing(df, excel_template_file_path, form_title, form_note, excelOutFileName):

    # Excel Template Setting
    column_to_find_last_row = ['B','C']
    last_column = 'K'
    max_char_one_line = 93

    # Obtain cell_coordinates of KEY_WORDS
    wb, ws, key_coord_dict = open_and_get_worksheet_key(excelTemplateFilePath=excel_template_file_path)

    drug_count = len(df)
    if drug_count > 14:
        print('Max 14 Drugs. Program Exit')
        exit()

    # SET FORM INFO
    set_form_title(ws, key_coord_dict['form_title_address'], form_title)
    set_request_date_time(ws, key_coord_dict['time_stamp_address'])

    # set_patient_info(ws, key_coord_dict['form_title_address'],window)
    # set_recipient_info(ws, key_coord_dict['form_title_address'],window)
    set_approve_title(ws,key_coord_dict['drug_begin_address'],last_column)
    set_form_note(ws, key_coord_dict['form_note_address'], form_note=thank_you_note_checker(message=form_note))
    transfer_df_to_drug_table(ws,df,key_coord_dict,last_column,max_char_one_line)
    set_signature_date_section(ws,df,key_coord_dict['signature_key_address'])

    last_row = find_last_row(ws,col_to_find_last_row=column_to_find_last_row)  # MUST PLACE RIGHT BEFORE CONFIDENTIAL NOTE

    set_confidential_note(ws, key_coord_dict['confidential_notice_address'], last_column)

    hide_extra_row(ws=ws, last_row=last_row,confidential_row = ws[key_coord_dict['confidential_notice_address']].row)

    adjust_padding_row_height(ws, drug_count=drug_count,confidential_coord=key_coord_dict['confidential_notice_address'])

    wb.save(excelOutFileName)

if __name__ == '__main__':
    drug_info_df = create_df_for_testing(drug_count = 2)
    drug_name_col_header = 'DRUG NAME'
    quantity_request_col_header = 'QTY DSP'
    refill_request_col_header = 'RF'
    sig_direction_col_header = 'Direction'

    drug_info_df['drug_top_line_info'] = drug_info_df[drug_name_col_header] + ' #' + drug_info_df[quantity_request_col_header].astype(str) \
                                         + ' --- Refill: ' + drug_info_df[refill_request_col_header].astype(str)

    # print(df[[drug_name_col_header, 'drug_top_line_info']])
    # ---------------------------------------------------------------------------------------------------------------------------

    specific_med_transfer_note = 'Per patient request, please transfer the following prescription(s) to HelloRx Pharmacy.'
    profile_transfer_note = 'Per patient request, please transfer the following prescription(s) to HelloRx Pharmacy.'
    refill_request_note = 'Per patient request, please authorize the following prescription(s) below.'

    empty_df = pd.DataFrame()

    processing(df=drug_info_df,
               excel_template_file_path='RxRequestTemplate.xlsx',
               form_title='Transfer Out Request',
               form_note=specific_med_transfer_note,
               excelOutFileName='transferRequest_temp.xlsx')

    convert_to_pdf(excelFilePath=r"D:\__HelloRx Pharmacy\GoogleDriveMirror\Python\HelloRxPharmacy\EzRx\transferRequest_temp.xlsx",
                   pdfOutFilePath=r"D:\__HelloRx Pharmacy\GoogleDriveMirror\Python\HelloRxPharmacy\EzRx\transferRequest_temp.pdf")


    # os.startfile(r"D:\__HelloRx Pharmacy\GoogleDriveMirror\Python\HelloRxPharmacy\EzRx\transferRequest_temp.pdf", "print")

